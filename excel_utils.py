import openpyxl

TITLE_CELL = 1
AUTHOR_CELL = 2


def sheet_writer(book, sheet_loc):
    """
    receives a book class and an excel worksheet and adds it in, adding the id number for the book
    :param book: book class
    :param sheet_loc: location of the worksheet
    """
    wb = openpyxl.load_workbook(sheet_loc)
    ws = wb.active
    all_rows = tuple(ws.rows)
    if not is_dup_book(book, all_rows):
        new_row_num = len(all_rows) + 1
        ws.cell(row=new_row_num, column=1, value=new_row_num - 1)
        book_tup = book.to_tuple()
        for x in range(7):
            ws.cell(new_row_num, column=x+2, value=book_tup[x])
        print(f"{book.title} by {book.author} has been added")
    else:
        print(f"{book.title} by {book.author} already exists in the database")
    wb.save(sheet_loc)


def is_dup_book(book, books_tuple):
    """
    receives an excel sheet as row tuples and a book and checks if the book is already in the sheet
    :param book: a book class
    :param books_tuple: a tuple of all rows in the sheet
    :return: True if is duplicate, otherwise False
    """
    for book_row in books_tuple:
        if book_row[TITLE_CELL].value == book.title and book_row[AUTHOR_CELL].value == book.author:
            return True
    return False
